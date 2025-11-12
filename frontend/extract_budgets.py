#!/usr/bin/env python3
"""
Extract crop budgets from Excel file and convert to JSON format
"""
import openpyxl
import json
import re

def clean_value(value):
    """Clean and convert cell values"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()

def extract_budget_from_sheet(sheet):
    """Extract budget data from a single sheet"""
    budget = {
        'crop_name': '',
        'summary': {},
        'costs': {
            'land_prep': [],
            'seed': [],
            'fertilizer': [],
            'chemicals': [],
            'labour': [],
            'irrigation': [],
            'transport_in': [],
            'transport_out': [],
            'sundry': []
        },
        'raw_data': []
    }
    
    # Get crop name from row 2
    crop_name = clean_value(sheet.cell(2, 1).value)
    if crop_name:
        budget['crop_name'] = crop_name
    
    # Extract summary data - detect sheet format
    # Some sheets have "GROSS MARGIN SUMMARY" header at row 4, data starts row 5
    # Other sheets start data directly at row 4
    row4_col2 = clean_value(sheet.cell(4, 2).value)
    
    # Check if row 4 is a header or data
    if row4_col2 and 'GROSS MARGIN SUMMARY' in row4_col2.upper():
        # Format 1: Header at row 4, data starts row 5
        start_row = 5
    else:
        # Format 2: Data starts at row 4
        start_row = 4
    
    summary_fields = [
        'gross_yield_kg_per_ha',
        'pack_out_percent',
        'net_yield_kg_per_ha',
        'variable_costs_per_ha',
        'variable_costs_per_kg',
        'farm_gate_price',
        'gross_return',
        'gross_profit',
        'return_per_dollar'
    ]
    
    for i, field_name in enumerate(summary_fields):
        row_num = start_row + i
        value = sheet.cell(row_num, 5).value  # Column 5 (E) has the values
        if value is not None:
            budget['summary'][field_name] = value
    
    # Extract cost breakdown (starting from row 16)
    current_category = None
    
    for row_idx in range(17, 100):  # Scan up to row 100
        col1 = clean_value(sheet.cell(row_idx, 1).value)
        col2 = clean_value(sheet.cell(row_idx, 2).value)
        col3 = clean_value(sheet.cell(row_idx, 3).value)
        col4 = clean_value(sheet.cell(row_idx, 4).value)
        col5 = clean_value(sheet.cell(row_idx, 5).value)
        col6 = clean_value(sheet.cell(row_idx, 6).value)
        col7 = clean_value(sheet.cell(row_idx, 7).value)
        
        # Check if this is a category header (col2 has category, col3 has first item)
        if col2 and col2.isupper() and col2 != 'DESCRIPTION':
            category_name = col2.lower().replace('-', '_').replace(' ', '_').strip()
            if 'land' in category_name or 'prep' in category_name:
                current_category = 'land_prep'
            elif 'seed' in category_name:
                current_category = 'seed'
            elif 'fertilizer' in category_name:
                current_category = 'fertilizer'
            elif 'chemical' in category_name:
                current_category = 'chemicals'
            elif 'labour' in category_name:
                current_category = 'labour'
            elif 'irrigation' in category_name:
                current_category = 'irrigation'
            elif 'transport' in category_name and 'in' in category_name:
                current_category = 'transport_in'
            elif 'transport' in category_name and 'out' in category_name:
                current_category = 'transport_out'
            elif 'sundry' in category_name or 'miscellaneous' in category_name:
                current_category = 'sundry'
            # Don't continue - the first item is on the same row!
        
        # Check if this is a cost item
        if col3 and current_category:
            item = {
                'description': col3,
                'diesel_lts': col4,
                'quantity': col5,
                'unit': col6,
                'unit_cost': col7
            }
            
            # Calculate total cost
            try:
                qty = float(col5) if col5 else 0
                unit_cost = float(col7) if col7 else 0
                diesel = float(col4) if col4 else 0
                
                if diesel > 0 and unit_cost > 0:
                    item['total_cost'] = diesel * unit_cost
                elif qty > 0 and unit_cost > 0:
                    item['total_cost'] = qty * unit_cost
                else:
                    item['total_cost'] = col4 if col4 else 0
            except:
                item['total_cost'] = 0
            
            budget['costs'][current_category].append(item)
        
        # Stop if we hit the total or disclaimer
        if col2 and ('TOTAL COST' in col2 or 'Note :' in col2 or 'Disclaimer' in col2):
            break
    
    return budget

def main():
    """Main extraction function"""
    print("Loading Excel file...")
    wb = openpyxl.load_workbook('/Users/providencemtendereki/Desktop/Crop Budgets.xlsx', data_only=True)
    
    all_budgets = {}
    
    # Skip 'Sheet1' as it's likely empty
    sheets_to_process = [name for name in wb.sheetnames if name != 'Sheet1']
    
    print(f"\nProcessing {len(sheets_to_process)} crop budget sheets...")
    
    for sheet_name in sheets_to_process:
        print(f"  - Processing: {sheet_name}")
        sheet = wb[sheet_name]
        
        try:
            budget_data = extract_budget_from_sheet(sheet)
            
            # Use sheet name as key if crop_name is empty
            crop_key = budget_data['crop_name'] if budget_data['crop_name'] else sheet_name
            
            all_budgets[crop_key] = budget_data
            
        except Exception as e:
            print(f"    ERROR processing {sheet_name}: {e}")
            continue
    
    wb.close()
    
    # Save to JSON file
    output_file = 'crop_budgets_data.json'
    print(f"\nSaving to {output_file}...")
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_budgets, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Successfully extracted {len(all_budgets)} crop budgets!")
    print(f"📄 Output saved to: {output_file}")
    
    # Print summary
    print("\n" + "="*80)
    print("BUDGET SUMMARY:")
    print("="*80)
    
    for crop_name, budget in all_budgets.items():
        summary = budget.get('summary', {})
        gross_margin = summary.get('gross_profit', 'N/A')
        gross_return = summary.get('gross_return', 'N/A')
        costs = summary.get('variable_costs_per_ha', 'N/A')
        
        print(f"\n{crop_name}:")
        print(f"  Gross Return: ${gross_return}")
        print(f"  Variable Costs: ${costs}")
        print(f"  Gross Margin: ${gross_margin}")

if __name__ == '__main__':
    main()
